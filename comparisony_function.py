import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import numpy as np

# ======================
# 配置参数
# ======================
INPUT_FILE = '/hpcfs/fhome/wupeng/wp/memote/genome_proteins/batch_analysis.xlsx'
SHEET_NAME = 'BIGG_model'
OUTPUT_FILE = "bigg模型比较_分析结果.xlsx"
CHART_PATH = "model_comparison_0.2.png"
CHART_svg = "model_comparison_0.2.svg"

# ======================
# 数据加载与预处理
# ======================
def load_and_process():
    """加载并处理数据"""
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
    
    # 清洗数据
    df = df.dropna(how='all').reset_index(drop=True)
    
    # 定义增强型分类逻辑
    def classify_model(name):
        name = str(name)
        if '.faa' in name:
            return 'CarveMe'
        if '.seed' in name:
            return 'ModelSEED'
        if 'GCF' in name and not any(x in name for x in ['.faa', '.seed']):
            return 'PEAR'
        if any(part.startswith('i') and part[1:].isalnum() 
              for part in name.split('_')):
            return 'Published'
        return 'Other'
    
    df['model_type'] = df['fasta_name'].apply(classify_model)
    return df[df['model_type'].isin(['PEAR', 'CarveMe', 'ModelSEED', 'Published'])]

# ======================
# 可视化函数
# ======================
def create_visualization(df):
    """创建增强的可视化图表"""
    plt.figure(figsize=(18, 12))
    metrics = ['nadh', 'atp', 'biomass', 'reactions', 'metabolites', 'genes']
    palette = {'PEAR':'#1f77b4', 'CarveMe':'#ff7f0e', 
              'ModelSEED':'#2ca02c', 'Published':'#d62728'}
    
    # 数据标准化
    df_log = df.copy()
    # for col in metrics:
    #     if col in ['nadh', 'atp', 'biomass']:  # 仅这三列取log
    #         df_log[col] = np.log1p(df[col])
    #     else:  # 其他列保持原值
    #         df_log[col] = df[col]
    
    # 创建分面图
    g = sns.FacetGrid(df_log.melt(id_vars=['model_type'], value_vars=metrics),
                     col='variable', col_wrap=3, sharey=False, height=4)
    g.map_dataframe(sns.violinplot, x='model_type', y='value',  bw=0.2,
                   palette=palette, inner="quartile")
    
    # 设置坐标轴标签
    for ax in g.axes.flat:
        metric = ax.get_title().split('=')[-1].strip()
        ax.set_title(metric, fontsize=12)
        ax.set_xlabel('')

        ax.set_ylabel('product rate')  # 其他仍使用对数刻度
        # 修改 y 轴标签
        if metric == 'reactions':
            ax.set_ylabel('Number of reactions')  # 直接显示数值
        if metric == 'metabolites':
            ax.set_ylabel('Number of metabolites')  # 直接显示数值
        if metric == 'genes':
            ax.set_ylabel('Number of genes')  # 直接显示数值

        ax.tick_params(axis='x', rotation=45)
    
    plt.tight_layout()
    plt.savefig(CHART_PATH, dpi=300, bbox_inches='tight')
    plt.savefig(CHART_svg)
    plt.show()
    plt.close()

# ======================
# 保存结果到Excel
# ======================
def save_to_excel(df):
    """保存结果到新Excel文件"""
    # 创建统计摘要
    stats = df.groupby('model_type').agg({
        'nadh': ['mean', 'median'],
        'atp': ['mean', 'median'],
        'biomass': ['mean', 'median'],
        'reaction': ['mean', 'median'],
        'metabolites': ['mean', 'median'],
        'gene': ['mean', 'median']
    })
    
    # 保存到Excel（写入数据和图表）
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='原始数据', index=False)
        stats.to_excel(writer, sheet_name='统计摘要')
        
        # 添加图表到一个新工作表
        wb = writer.book
        ws = wb.create_sheet('可视化')
        img = Image(CHART_PATH)
        ws.add_image(img, 'A1')
    
    # 调整列宽
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(OUTPUT_FILE)
    
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        if sheetname == '可视化':
            ws.column_dimensions['A'].width = 50
        else:
            # 遍历所有列（通过列索引方式）
            for col in range(1, ws.max_column + 1):
                max_length = 0
                # 遍历每一列的所有单元格
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(OUTPUT_FILE)

# ======================
# 主程序
# ======================
if __name__ == "__main__":
    try:
        # 执行分析
        df = load_and_process()
        create_visualization(df)
        save_to_excel(df)
        
        print(f"分析完成！结果已保存至：{OUTPUT_FILE}")
        print("包含以下工作表：")
        print("1. 原始数据 - 清洗后的原始数据")
        print("2. 统计摘要 - 各指标统计值")
        print("3. 可视化 - 模型比较图表")
        
    except Exception as e:
        print(f"执行出错：{str(e)}")
        print("请检查：")
        print("1. Excel文件路径是否正确")
        print("2. 工作表名称是否匹配")
        print("3. 是否安装所需依赖库（pandas, openpyxl, seaborn）")
